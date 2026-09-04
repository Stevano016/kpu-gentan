# -*- coding: utf-8 -*-
"""
Membangun `database/seeders/dpt_seed.csv` dari berkas Excel DPS — baik satu
berkas gabungan (sheet `RW 1` .. `RW 14`) maupun direktori `RW *.xlsx`.

Pemakaian:
    python bangun_dpt_seed.py "<berkas.xlsx>" [--tanpa-db]

Secara baku data baru dicocokkan dengan database lokal supaya NIK/NKK yang
sudah ada tidak hilang ketika berkas sumbernya belum lengkap. `--tanpa-db`
mematikan pencocokan itu, dipakai ketika berkas sumber memang dimaksudkan
menggantikan seluruh data seeder: hasilnya lalu bisa dibangun ulang dari
berkas itu saja, tidak bergantung pada isi database saat ini.

Nomor urut asli dari berkas Excel tetap dicatat pada kolom `no_urut`.
"""

import csv
import sys
import os
import glob
import sqlite3
from pathlib import Path
import openpyxl

AKAR = Path(__file__).resolve().parents[3].parent  # .../KPPS Gentan
DB_PATH = Path(__file__).resolve().parents[2] / "database.sqlite"
TUJUAN = Path(__file__).resolve().parents[1] / "dpt_seed.csv"

AWALAN_NIK_SINTETIS = "9999"
AWALAN_NKK_SINTETIS = "9998"

# Pembagian wilayah TPS, disalin dari surat pembagian TPS Kelurahan Gentan.
PETA_TPS = {
    1: {"001": None, "002": None, "010": ["006", "007"]},
    2: {"003": None, "004": None, "014": None, "006": ["002", "004", "006", "008"]},
    3: {"007": None, "013": None, "006": ["001", "003", "005", "007"], "009": ["001"]},
    4: {"008": None, "012": None},
    5: {"005": None, "011": None,
        "009": ["002", "003", "004", "005"],
        "010": ["001", "002", "003", "004", "005"]},
}
TPS_CADANGAN = 1

KOLOM_KELUARAN = [
    "nik", "nkk", "nama", "tps_id", "jenis_kelamin", "umur", "status_kawin",
    "alamat", "rt", "rw", "pekerjaan", "disabilitas", "catatan_impor",
    "nik_sintetis", "nkk_sintetis", "no_urut"
]


def hitung_umur_dan_gender(nik: str):
    if not nik or len(nik) != 16 or not nik.isdigit() or nik.startswith(AWALAN_NIK_SINTETIS) or nik.startswith(AWALAN_NKK_SINTETIS):
        return None, None
        
    try:
        day_str = nik[6:8]
        month_str = nik[8:10]
        year_str = nik[10:12]
        
        day_val = int(day_str)
        month_val = int(month_str)
        year_val = int(year_str)
        
        if 1 <= month_val <= 12:
            is_female = False
            day = day_val
            if day > 40:
                is_female = True
                day = day - 40
                
            if 1 <= day <= 31:
                # yearVal <= 9 ? 2000 + yearVal : 1900 + yearVal
                birth_year = 2000 + year_val if year_val <= 9 else 1900 + year_val
                
                from datetime import date
                today = date(2026, 8, 27)
                birth_date = date(birth_year, month_val, day)
                
                age = today.year - birth_date.year
                if (today.month, today.day) < (birth_date.month, birth_date.day):
                    age -= 1
                    
                gender = "PEREMPUAN" if is_female else "LAKI-LAKI"
                return age, gender
    except Exception:
        pass
    return None, None


def teks(nilai) -> str:
    if nilai is None:
        return ""
    if isinstance(nilai, float) and nilai.is_integer():
        nilai = int(nilai)
    return str(nilai).strip()


def nomor_wilayah(nilai: str) -> str:
    nilai = teks(nilai)
    if not nilai or nilai == "-":
        return ""
    return nilai.lstrip("0").zfill(3) if nilai.isdigit() else nilai


def cari_tps(rw: str, rt: str) -> int:
    for tps_id, wilayah in PETA_TPS.items():
        if rw in wilayah:
            daftar_rt = wilayah[rw]
            if daftar_rt is None or rt in daftar_rt:
                return tps_id
    return TPS_CADANGAN


def muat_data_db():
    """Memuat data DPT dari SQLite lokal jika ada."""
    dpt_map_full = {}
    dpt_map_name = {}
    if not DB_PATH.exists():
        print(f"Database tidak ditemukan di {DB_PATH}, pencocokan dilewati.")
        return dpt_map_full, dpt_map_name

    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        # Cek apakah tabel dpt ada
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='dpt'")
        if not cursor.fetchone():
            conn.close()
            return dpt_map_full, dpt_map_name

        cursor.execute("SELECT nik, nkk, nama, rt, rw, nik_sintetis, nkk_sintetis FROM dpt")
        for row in cursor.fetchall():
            nik, nkk, nama, rt, rw, nik_sint, nkk_sint = row
            nama_clean = nama.strip().upper()
            rt_clean = nomor_wilayah(rt)
            rw_clean = nomor_wilayah(rw)
            
            val = {
                "nik": nik,
                "nkk": nkk,
                "nik_sintetis": bool(nik_sint),
                "nkk_sintetis": bool(nkk_sint)
            }
            # Map by full (nama, rt, rw)
            dpt_map_full[(nama_clean, rt_clean, rw_clean)] = val
            # Map by name
            dpt_map_name.setdefault(nama_clean, []).append({
                "nik": nik,
                "nkk": nkk,
                "rt": rt_clean,
                "rw": rw_clean,
                "nik_sintetis": bool(nik_sint),
                "nkk_sintetis": bool(nkk_sint)
            })
        conn.close()
        print(f"Berhasil memuat {len(dpt_map_full)} data dari database untuk pencocokan.")
    except Exception as e:
        print(f"Gagal membaca database: {e}")
    return dpt_map_full, dpt_map_name


def main() -> int:
    argumen = [a for a in sys.argv[1:] if not a.startswith("--")]
    tanpa_db = "--tanpa-db" in sys.argv[1:]

    excel_path = Path(argumen[0]) if argumen else AKAR / "DPS_PILKADES_GENTAN_2026_GABUNGAN_BERSIH (2).xlsx"
    if not excel_path.exists():
        excel_path = AKAR / "DPS PILKADES 2026"
        if not excel_path.exists():
            print(f"Berkas/Direktori data tidak ditemukan: {excel_path}")
            return 1

    is_single_file = excel_path.is_file()
    
    if is_single_file:
        print(f"Menggunakan berkas gabungan: {excel_path.name}")
        wb_global = openpyxl.load_workbook(excel_path, data_only=True)
    else:
        print(f"Menggunakan direktori berkas: {excel_path}")
        berkas_list = sorted(
            glob.glob(os.path.join(excel_path, "RW *.xlsx")),
            key=lambda x: int(os.path.basename(x).replace("RW ", "").replace(".xlsx", ""))
        )
        if not berkas_list:
            print(f"Tidak ditemukan berkas RW *.xlsx di {excel_path}")
            return 1

    if tanpa_db:
        print("Mode --tanpa-db: seluruh NIK/NKK diambil dari berkas Excel, database diabaikan.")
        db_full, db_name = {}, {}
    else:
        db_full, db_name = muat_data_db()

    hasil = []
    nik_terpakai = set()
    nkk_terpakai = set()

    # Hitung pembuat nomor sementara untuk data baru yang tidak cocok
    # dan tidak memiliki NIK/NKK di berkas Excel.
    urut_nik_sintetis = 0
    urut_nkk_sintetis = 0
    
    # Simpan NIK/NKK dari DB untuk pengecekan konflik saat pembuatan nomor sementara baru
    db_niks = set()
    db_nkks = set()
    for key, val in db_full.items():
        if val["nik"]:
            db_niks.add(val["nik"])
            if val["nik"].startswith(AWALAN_NIK_SINTETIS):
                try:
                    num = int(val["nik"][len(AWALAN_NIK_SINTETIS):])
                    if num > urut_nik_sintetis:
                        urut_nik_sintetis = num
                except ValueError:
                    pass
        if val["nkk"]:
            db_nkks.add(val["nkk"])
            if val["nkk"].startswith(AWALAN_NKK_SINTETIS):
                try:
                    num = int(val["nkk"][len(AWALAN_NKK_SINTETIS):])
                    if num > urut_nkk_sintetis:
                        urut_nkk_sintetis = num
                except ValueError:
                    pass

    # nik_terpakai and nkk_terpakai will track numbers generated/used in this run
    nik_terpakai = set()
    nkk_terpakai = set()

    global_no_urut = 0
    jumlah = {
        "total": 0,
        "cocok_full": 0,
        "cocok_nama": 0,
        "baru_dari_excel": 0,
        "nik_sintetis": 0,
        "nkk_sintetis": 0,
        "nik_kembar": 0
    }
    # Nomor identitas yang panjangnya bukan 16 digit; dikumpulkan untuk
    # dilaporkan di akhir, bukan didiamkan di tengah keluaran yang panjang.
    cacat = []

    for rw_idx in range(1, 15):
        rw_num = str(rw_idx)
        sheet_name = f"RW {rw_idx}"
        
        if is_single_file:
            if sheet_name not in wb_global.sheetnames:
                print(f"  Warning: sheet {sheet_name} tidak ditemukan. Dilewati.")
                continue
            ws = wb_global[sheet_name]
            fname = excel_path.name
        else:
            filepath = None
            for f in berkas_list:
                basename = os.path.basename(f)
                if f"RW {rw_idx}.xlsx" == basename or f"RW {str(rw_idx).zfill(2)}.xlsx" == basename or f"RW {str(rw_idx).zfill(3)}.xlsx" == basename:
                    filepath = f
                    break
            if not filepath:
                for f in berkas_list:
                    basename = os.path.basename(f)
                    if f"RW {rw_idx}" in basename or f"RW {str(rw_idx).zfill(2)}" in basename:
                        filepath = f
                        break
            if not filepath:
                print(f"  Warning: Berkas untuk RW {rw_idx} tidak ditemukan. Dilewati.")
                continue
            fname = os.path.basename(filepath)
            wb = openpyxl.load_workbook(filepath, data_only=True)
            if sheet_name not in wb.sheetnames:
                found_sheet = None
                for s in wb.sheetnames:
                    if s.strip().upper() == sheet_name.upper() or s.strip().upper() == f"RW {str(rw_idx).zfill(2)}".upper():
                        found_sheet = s
                        break
                if not found_sheet:
                    print(f"  Warning: sheet {sheet_name} tidak ditemukan di {fname}. Dilewati.")
                    continue
                ws = wb[found_sheet]
            else:
                ws = wb[sheet_name]
                
        print(f"Memproses {fname} sheet {sheet_name}...")
        rows = list(ws.iter_rows(values_only=True))
        
        header_row = -1
        for r_idx, r in enumerate(rows):
            if r and any(str(c).strip().upper() == "NAMA LENGKAP" for c in r if c is not None):
                header_row = r_idx
                break
                
        if header_row == -1:
            print(f"  Warning: header 'NAMA LENGKAP' tidak ditemukan. Dilewati.")
            continue
            
        headers = [str(c).strip().upper() if c is not None else "" for c in rows[header_row]]
        name_idx = headers.index("NAMA LENGKAP")
        
        ktp_idx = -1
        for name in ["NOMOR KTP (NIK)", "NOMOR KTP", "NIK"]:
            if name in headers:
                ktp_idx = headers.index(name)
                break
                
        kk_idx = -1
        for name in ["NOMOR KK (NKK)", "NOMOR KK", "NKK"]:
            if name in headers:
                kk_idx = headers.index(name)
                break
                
        rt_idx = headers.index("RT") if "RT" in headers else -1
        rw_idx = headers.index("RW") if "RW" in headers else -1
        alamat_idx = headers.index("ALAMAT") if "ALAMAT" in headers else -1
        umur_idx = headers.index("UMUR") if "UMUR" in headers else -1
        kawin_idx = headers.index("KAWIN/SUDAH PERNAH KAWIN/BELUM") if "KAWIN/SUDAH PERNAH KAWIN/BELUM" in headers else -1
        gender_idx = headers.index("L/P") if "L/P" in headers else -1
        pekerjaan_idx = headers.index("PEKERJAAN") if "PEKERJAAN" in headers else -1
        disabilitas_idx = headers.index("DISABILITAS") if "DISABILITAS" in headers else -1
        ket_idx = headers.index("KETERANGAN") if "KETERANGAN" in headers else -1
        
        for r_idx, r in enumerate(rows[header_row+1:], start=header_row+2):
            if len(r) > name_idx and r[name_idx]:
                name_val = teks(r[name_idx]).strip().upper()
                # Skip header/column number rows
                if not name_val or name_val in ("NAMA LENGKAP", "NO", "ORI") or name_val.isdigit() or len(name_val) <= 2:
                    continue
                if teks(r[0]) == "ORI" or (len(r) > 1 and teks(r[1]) == "ORI"):
                    continue
                
                rt = nomor_wilayah(r[rt_idx]) if rt_idx != -1 and len(r) > rt_idx else ""
                rw = nomor_wilayah(r[rw_idx]) if rw_idx != -1 and len(r) > rw_idx else ""
                excel_nik = teks(r[ktp_idx]) if ktp_idx != -1 and len(r) > ktp_idx else ""
                excel_nik = excel_nik.replace(" ", "").replace("-", "")
                # Nomor cacat panjang dikosongkan, bukan menggugurkan proses.
                #
                # Dulu satu digit yang kurang menggagalkan pembangunan seluruh
                # berkas seeder, jadi tujuh ribu baris lain ikut tertahan oleh
                # satu salah ketik. Sekarang orangnya tetap masuk daftar dengan
                # nomor sementara — ia memang pemilih — dan angka aslinya
                # disimpan di kolom catatan supaya bisa dibetulkan, bukan
                # ditebak di sini. Semuanya dilaporkan di akhir keluaran.
                if excel_nik and (len(excel_nik) != 16 or not excel_nik.isdigit()):
                    cacat.append((sheet_name, r_idx, name_val, "NIK", excel_nik))
                    catatan_cacat_nik = (
                        f"NIK pada berkas sumber tidak sah ('{excel_nik}', "
                        f"{len(excel_nik)} digit) — wajib diperbaiki."
                    )
                    excel_nik = ""
                else:
                    catatan_cacat_nik = ""

                excel_nkk = teks(r[kk_idx]) if kk_idx != -1 and len(r) > kk_idx else ""
                excel_nkk = excel_nkk.replace(" ", "").replace("-", "")
                if excel_nkk and (len(excel_nkk) != 16 or not excel_nkk.isdigit()):
                    cacat.append((sheet_name, r_idx, name_val, "NKK", excel_nkk))
                    catatan_cacat_nkk = (
                        f"NKK pada berkas sumber tidak sah ('{excel_nkk}', "
                        f"{len(excel_nkk)} digit) — wajib diperbaiki."
                    )
                    excel_nkk = ""
                else:
                    catatan_cacat_nkk = ""

                catatan = [c for c in [teks(r[ket_idx]) if ket_idx != -1 and len(r) > ket_idx else ""] if c]
                catatan += [c for c in (catatan_cacat_nik, catatan_cacat_nkk) if c]

                # Tentukan NIK/NKK dan apakah sintetis
                nik = ""
                nkk = ""
                nik_sintetis = False
                nkk_sintetis = False
                
                # Strategi Pencocokan Database
                db_match = None
                is_cocok_full = False
                is_cocok_nama = False
                
                # 1. Nama + RT + RW
                if (name_val, rt, rw) in db_full:
                    db_match = db_full[(name_val, rt, rw)]
                    is_cocok_full = True
                # 2. Nama saja (jika unik di DB)
                elif name_val in db_name and len(db_name[name_val]) == 1:
                    db_match = db_name[name_val][0]
                    is_cocok_nama = True
                
                if db_match:
                    # Memprioritaskan NIK/NKK dari Excel jika ia merupakan NIK/NKK riil (bukan sintetis/kosong)
                    if excel_nik and not excel_nik.startswith(AWALAN_NIK_SINTETIS):
                        nik = excel_nik
                        nik_sintetis = False
                    else:
                        nik = db_match["nik"]
                        nik_sintetis = db_match["nik_sintetis"]
                        
                    if excel_nkk and not excel_nkk.startswith(AWALAN_NKK_SINTETIS):
                        nkk = excel_nkk
                        nkk_sintetis = False
                    else:
                        nkk = db_match["nkk"]
                        nkk_sintetis = db_match["nkk_sintetis"]
                else:
                    # Data Baru dari Excel
                    if excel_nik:
                        nik = excel_nik
                        nik_sintetis = excel_nik.startswith(AWALAN_NIK_SINTETIS)
                    if excel_nkk:
                        nkk = excel_nkk
                        nkk_sintetis = excel_nkk.startswith(AWALAN_NKK_SINTETIS)

                # Tangani jika NIK/NKK kosong
                if not nik:
                    while True:
                        urut_nik_sintetis += 1
                        nik = AWALAN_NIK_SINTETIS + str(urut_nik_sintetis).zfill(12)
                        if nik not in db_niks and nik not in nik_terpakai:
                            break
                    nik_sintetis = True
                    catatan.append("NIK belum ada di data pembanding — nomor sementara, wajib dilengkapi pantarlih.")
                
                if not nkk:
                    while True:
                        urut_nkk_sintetis += 1
                        nkk = AWALAN_NKK_SINTETIS + str(urut_nkk_sintetis).zfill(12)
                        if nkk not in db_nkks and nkk not in nkk_terpakai:
                            break
                    nkk_sintetis = True
                    catatan.append("NKK belum ada di data pembanding — nomor sementara, keluarganya belum bisa dikelompokkan.")

                # Tangani duplikat NIK di hasil penulisan baru (Hanya seed 1 NIK 1 Nama 1 NKK)
                if nik in nik_terpakai:
                    print(f"  Warning: Voter '{name_val}' (Excel baris {r_idx}) dilewati karena NIK '{nik}' duplikat.")
                    continue

                # Setelah dipastikan tidak diskip, baru update counter dan record
                global_no_urut += 1
                jumlah["total"] += 1

                if db_match:
                    if is_cocok_full:
                        jumlah["cocok_full"] += 1
                    elif is_cocok_nama:
                        jumlah["cocok_nama"] += 1
                else:
                    jumlah["baru_dari_excel"] += 1

                nik_terpakai.add(nik)
                nkk_terpakai.add(nkk)

                if nik_sintetis:
                    jumlah["nik_sintetis"] += 1
                if nkk_sintetis:
                    jumlah["nkk_sintetis"] += 1

                # Kalkulasi umur dan gender secara deterministik
                calc_umur, calc_gender = hitung_umur_dan_gender(nik)

                umur = teks(r[umur_idx]) if umur_idx != -1 and len(r) > umur_idx else ""
                gender = teks(r[gender_idx]) if gender_idx != -1 and len(r) > gender_idx else ""

                if calc_umur is not None:
                    umur = str(calc_umur)
                if calc_gender is not None:
                    gender = calc_gender

                hasil.append({
                    "nik": nik,
                    "nkk": nkk,
                    "nama": teks(r[name_idx]),
                    "tps_id": cari_tps(rw, rt),
                    "jenis_kelamin": gender,
                    "umur": umur if umur.isdigit() else "",
                    "status_kawin": teks(r[kawin_idx]) if kawin_idx != -1 and len(r) > kawin_idx else "",
                    "alamat": teks(r[alamat_idx]) if alamat_idx != -1 and len(r) > alamat_idx else "",
                    "rt": rt,
                    "rw": rw,
                    "pekerjaan": teks(r[pekerjaan_idx]) if pekerjaan_idx != -1 and len(r) > pekerjaan_idx else "",
                    "disabilitas": teks(r[disabilitas_idx]) if disabilitas_idx != -1 and len(r) > disabilitas_idx else "",
                    "catatan_impor": " | ".join(catatan),
                    "nik_sintetis": "1" if nik_sintetis else "0",
                    "nkk_sintetis": "1" if nkk_sintetis else "0",
                    "no_urut": global_no_urut
                })

    # Urutan CSV dibuat menetap per TPS/RW/RT/nama untuk kenyamanan database/seeder,
    # tetapi `no_urut` yang tersimpan di baris akan melacak nomor urutan asli Excel.
    # Namun, agar id_pemilih (dari seeder) mengikuti urutan per TPS/RW/RT/nama secara teratur,
    # kita tetap mengurutkan CSV seperti semula.
    hasil.sort(key=lambda b: (b["tps_id"], b["rw"], b["rt"], b["nama"], b["nik"]))

    with TUJUAN.open("w", encoding="utf-8", newline="") as f:
        penulis = csv.DictWriter(f, fieldnames=KOLOM_KELUARAN)
        penulis.writeheader()
        penulis.writerows(hasil)

    print(f"\nSelesai! Ditulis ke {TUJUAN}")
    print(f"  Total baris          : {jumlah['total']}")
    print(f"  Cocok (Nama+RT+RW)   : {jumlah['cocok_full']}")
    print(f"  Cocok (Nama saja)    : {jumlah['cocok_nama']}")
    print(f"  Data baru            : {jumlah['baru_dari_excel']}")
    print(f"  NIK sementara        : {jumlah['nik_sintetis']} (termasuk {jumlah['nik_kembar']} duplikat)")
    print(f"  NKK sementara        : {jumlah['nkk_sintetis']}")
    
    per_tps = {}
    for b in hasil:
        per_tps[b["tps_id"]] = per_tps.get(b["tps_id"], 0) + 1
    for t in sorted(per_tps):
        print(f"  TPS {t}                : {per_tps[t]}")

    if cacat:
        print(f"\nPERHATIAN — {len(cacat)} nomor identitas pada berkas sumber bukan 16 digit.")
        print("  Barisnya tetap ikut memakai nomor sementara; angka aslinya ada di kolom catatan_impor.")
        for sheet, baris, nama, jenis, nilai in cacat:
            print(f"  {sheet} baris {baris}: {nama} — {jenis} '{nilai}' ({len(nilai)} digit)")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
